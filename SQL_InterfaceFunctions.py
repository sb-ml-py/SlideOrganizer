#Functions for interfacing with SQLite Database

#Import relevant Dictionaries
import sqlite3

def openDatabase(db_path):
    #Interface with database
    import sqlite3
    try:
        #Connect to DB and create a cursor
        con = sqlite3.connect(db_path)
        cur = con.cursor()
        print('DB Init')
        
        #Create table, slide, with categories ID_num, diagnosis, and tissue_type if not already created
        cur.execute("CREATE TABLE if not exists slide(case_ID, filename, stain, diagnosis, tissue_type, history, year)")
        
        return con, cur
    # Handle errors
    except sqlite3.Error as error:
        print('Error occurred - ', error)
#

def closeDatabase(con, cur):
    #Close interface with database
    if con:
        cur.close()
        con.close()
        print('Slidebox closed.')
#

def addEntry(con, cur, strline):
    #Add slide to slidebox
    #Insert Data into table
    cur.execute(strline)

    #Save (commit) the insertion
    con.commit()
#

def delEntry(con, cur, entry_ID):
    #Remove slide from slidebox
    delstr = "DELETE FROM slide WHERE filename = '" + str(entry_ID) + "';"
    cur.execute(delstr)
    #Save (commit) the insertion
    con.commit()
#

def modEntry(con, cur, Entry_name, ent_cat, ent_var):
    #Change details of slide
    modstr = "UPDATE slide SET " + ent_cat + " = '" + ent_var + "' WHERE filename = '" + Entry_name + "'"
    cur.execute(modstr)
    con.commit()
#

def addMany(con, cur, xlsxpath):
    import openpyxl
    wb = openpyxl.load_workbook(xlsxpath)
    sheet = wb.active

    for i in range(2, sheet.max_row):
        ID = str(sheet.cell(row = i, column = 1).value)
        name = str(sheet.cell(row = i, column = 2).value)
        stn = str(sheet.cell(row = i, column = 3).value)
        diag = str(sheet.cell(row = i, column = 4).value)
        tissue = str(sheet.cell(row = i, column = 5).value)
        hist = str(sheet.cell(row = i, column = 6).value)
        year = str(sheet.cell(row = i, column = 7).value)
        strline = "INSERT INTO slide VALUES ('" + ID + "', '" + name + "', '" + stn + "', '" + diag + "', '" + tissue + "', '" + hist + "', '" + year + "')"
        addEntry(con, cur, strline)
    #
#

def removeMany(con, cur, xlsxpath):
    import openpyxl
    wb = openpyxl.load_workbook(xlsxpath)
    sheet = wb.active

    for i in range(2, sheet.max_row):
        name = str(sheet.cell(row = i, column = 2).value)
        delEntry(con, cur, name)
    #
#

def printDB(table):
    import openpyxl
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'Slidebox'                #Name the sheet

    sheet.cell(row = 1, column = 1).value = 'Case ID'
    sheet.cell(row = 1, column = 2).value = 'Filename'
    sheet.cell(row = 1, column = 3).value = 'Stain'
    sheet.cell(row = 1, column = 4).value = 'Diagnosis'
    sheet.cell(row = 1, column = 5).value = 'Tissue'
    sheet.cell(row = 1, column = 6).value = 'History'
    sheet.cell(row = 1, column = 7).value = 'Year'

    for i in range(len(table)):
        for j in range(len(table[0])):
            sheet.cell(row = i+1, column = j+1).value = table[i][j]

    #Save the excel workbook
    wb.save('Slidebox_print.xlsx')
#