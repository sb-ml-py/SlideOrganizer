#Python SQLite Database Code

#Import relevant Dictionaries
import sqlite3
from tabulate import tabulate
import os

#Define functions
def openDatabase(db_path):
    #Interface with database
    import sqlite3
    try:
        #Connect to DB and create a cursor
        con = sqlite3.connect(db_path)
        cur = con.cursor()
        print('DB Init')
        
        #Create table, slide, with categories ID_num, diagnosis, and tissue_type if not already created
        cur.execute("CREATE TABLE if not exists slide(case_ID, filename, stain, diagnosis, tissue_type, history, year)")
        
        return con, cur
    # Handle errors
    except sqlite3.Error as error:
        print('Error occurred - ', error)
#

def closeDatabase(con, cur):
    #Close interface with database
    if con:
        cur.close()
        con.close()
        print('Slidebox closed.')
#

def addEntry(con, cur, strline):
    #Add slide to slidebox
    #Insert Data into table
    cur.execute(strline)

    #Save (commit) the insertion
    con.commit()
#

def delEntry(con, cur, entry_ID):
    #Remove slide from slidebox
    delstr = "DELETE FROM slide WHERE filename = '" + str(entry_ID) + "';"
    cur.execute(delstr)
    #Save (commit) the insertion
    con.commit()
#

def modEntry(con, cur, Entry_name, ent_cat, ent_var):
    #Change details of slide
    modstr = "UPDATE slide SET " + ent_cat + " = '" + ent_var + "' WHERE filename = '" + Entry_name + "'"
    cur.execute(modstr)
    con.commit()
#

def addMany(con, cur, xlsxpath):
    import openpyxl
    wb = openpyxl.load_workbook(xlsxpath)
    sheet = wb.active

    for i in range(2, sheet.max_row):
        ID = str(sheet.cell(row = i, column = 1).value)
        name = str(sheet.cell(row = i, column = 2).value)
        stn = str(sheet.cell(row = i, column = 3).value)
        diag = str(sheet.cell(row = i, column = 4).value)
        tissue = str(sheet.cell(row = i, column = 5).value)
        hist = str(sheet.cell(row = i, column = 6).value)
        year = str(sheet.cell(row = i, column = 7).value)
        strline = "INSERT INTO slide VALUES ('" + ID + "', '" + name + "', '" + stn + "', '" + diag + "', '" + tissue + "', '" + hist + "', '" + year + "')"
        addEntry(con, cur, strline)
    #
#

def removeMany(con, cur, xlsxpath):
    import openpyxl
    wb = openpyxl.load_workbook(xlsxpath)
    sheet = wb.active

    for i in range(2, sheet.max_row):
        name = str(sheet.cell(row = i, column = 2).value)
        delEntry(con, cur, name)
    #
#

def printDB(table, xlsxfilename):
    import openpyxl
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'Slidebox'                #Name the sheet

    sheet.cell(row = 1, column = 1).value = 'Case ID'
    sheet.cell(row = 1, column = 2).value = 'Filename'
    sheet.cell(row = 1, column = 3).value = 'Stain'
    sheet.cell(row = 1, column = 4).value = 'Diagnosis'
    sheet.cell(row = 1, column = 5).value = 'Tissue'
    sheet.cell(row = 1, column = 6).value = 'History'
    sheet.cell(row = 1, column = 7).value = 'Year'

    for i in range(len(table)):
        for j in range(len(table[0])):
            sheet.cell(row = i+1, column = j+1).value = table[i][j]

    #Save the excel workbook
    wb.save(xlsxfilename)
#

#Instructions for use:
print('''
Welcome to Slide Organizer

Options:
0 - Quit
1 - View all slides
2 - Search slides
3 - Add slide
4 - Remove slide
5 - Edit slide
6 - Add many slides, import from .xlsx file
7 - Remove many slides, specified by .xlsx file
8 - Print database to .xlsx file

Note: All entries are case-sensitive.
''')

#Open slidebox independent of drive location -- for use in removable media / USB drives
filepath = os.path.realpath(__file__)
folderpath = os.path.dirname(filepath)
os.chdir(folderpath)
cwd = os.getcwd()
dirpath = os.path.join(cwd, 'slidebox_files')
db_path = os.path.join(dirpath, 'slidebox.db')
if os.path.exists(db_path):
    con, cur = openDatabase(db_path)
    print('Directory exists')
else:
    os.makedirs(dirpath)
    con, cur = openDatabase(db_path)
    print('directory made')

try:
    while True:
        #Prompt action
        choice = input('Enter choice of action (1-8): ')
        #View slidebox
        if choice == '0':
            closeDatabase(con, cur)
            break
        elif choice == '1':
            res = cur.execute("SELECT case_ID, filename, stain, diagnosis, tissue_type, history, year FROM slide ORDER BY case_ID ASC;")
            reslist = res.fetchall()
            table = [('Case ID', 'Filename', 'Stain', 'Diagnosis', 'Tissue', 'History', 'Year')]
            for i in reslist:
                newtable = table.append(i)
            print(tabulate(table))
            #
        #Search slidebox
        elif choice == '2':
            print('''
            Search type:
            1 - Case ID
            2 - Slide name
            3 - Diagnosis
            4 - Tissue Type
            5 - Year
            
            Note: Queries are case-sensitive.
            ''')
            searchtype = input('Enter search type (1-5): ')
            i = True
            if searchtype == '1':
                stype = 'case_ID'
                searchterm = input('Enter Case ID: ')
            elif searchtype == '2':
                stype = 'filename'
                searchterm = input('Enter filename: ')
            elif searchtype == '3':
                stype = 'diagnosis'
                searchterm = input('Enter diagnosis: ')
            elif searchtype == '4':
                stype = 'tissue_type'
                searchterm = input('Enter tissue type: ')
            elif searchtype == '5':
                stype = 'year'
                searchterm = input('Enter year: ')
            else:
                print('Error: please select appropriate search type')
                i = False
            #
            if i:
                searchstr = str("SELECT case_ID, filename, stain, diagnosis, tissue_type, history, year FROM slide WHERE " + stype + " = '" + searchterm + "';")
                res = cur.execute(searchstr)
                reslist = res.fetchall()
                table = [('Case ID', 'Filename', 'Stain', 'Diagnosis', 'Tissue', 'History', 'Year')]
                for i in reslist:
                    table.append(i)
                print(tabulate(table))
                #
        #Add slide to slidebox
        elif choice == '3':
            ID = str(input('Enter Case ID: '))
            name = str(input('Enter filename: '))
            stn = str(input('Enter stain: '))
            diag = str(input('Enter diagnosis: '))
            tissue = str(input('Enter tissue type: '))
            hist = str(input('Enter history: '))
            year = str(input('Enter year: '))
            table = [('Case ID', 'Filename', 'Stain', 'Diagnosis', 'Tissue', 'History', 'Year'), (ID, name, stn, diag, tissue, hist, year)]
            print(tabulate(table))
            ns_conf = input('Is this correct? (Y/N): ')
            if ns_conf == 'Y':
                strline = "INSERT INTO slide VALUES ('" + ID + "', '" + name + "', '" + stn + "', '" + diag + "', '" + tissue + "', '" + hist + "', '" + year + "')"
                addEntry(con, cur, strline)
                print('Slide added.')
            else:
                print('Entry not confirmed. New slide not added.')
            #
        #Remove slide from slidebox
        elif choice == '4':
            entry_ID = input('Enter filename of slide to be deleted: ')
            delsld = "SELECT case_ID, filename, stain, diagnosis, tissue_type, history, year FROM slide WHERE filename = '" + entry_ID + "';"
            res = cur.execute(delsld)
            reslist = res.fetchall()
            table = [('Case ID', 'Filename', 'Stain', 'Diagnosis', 'Tissue', 'History', 'Year')]
            for i in reslist:
                newtable = table.append(i)
            print(tabulate(table))
            d_conf = input('Remove the above slide(s)? (Y/N): ')
            if d_conf == 'Y':
                delEntry(con, cur, entry_ID)
                print('Slide removed.')
            else:
                print('Deletion not confirmed. No slides removed.')
        #
        elif choice == '5':
            Entry_name = input('Enter filename of slide to by modified: ')
            print('''
            Search type:
            1 - Case ID
            2 - Slide name
            3 - Stain
            5 - Diagnosis
            5 - Tissue Type
            6 - History
            7 - Year
            
            Note: Entries are case-sensitive.
            ''')
            while True:
                ent_cat = input('Which information do you want to change? (1-7): ')
                if ent_cat == '1':
                    ent_cat = 'case_ID'
                    break
                elif ent_cat == '2':
                    ent_cat = 'name'
                    break
                elif ent_cat == '3':
                    ent_cat = 'stain'
                    break
                elif ent_cat == '4':
                    ent_cat = 'diagnosis'
                    break
                elif ent_cat == '5':
                    ent_cat = 'tissue_type'
                    break
                elif ent_cat == '6':
                    ent_cat = 'history'
                    break
                elif ent_cat == '7':
                    ent_cat = 'year'
                    break
            ent_var = input('Enter new information: ')
            mod_conf = input('For slide ' + Entry_name + ' change ' + ent_cat + ' to ' + ent_var + '.\nIs this correct? (Y/N): ')
            if mod_conf == 'Y':
                modEntry(con, cur, Entry_name, ent_cat, ent_var)
            else:
                print('Changes not confirmed. Changes not saved.')
        elif choice == '6':
            xlsxpath = input('Enter path of .xlsx file: ')
            if os.path.isabs(xlsxpath):

                import openpyxl
                wb = openpyxl.load_workbook(xlsxpath)
                sheet = wb.active

                table = [('Case ID', 'Filename', 'Stain', 'Diagnosis', 'Tissue', 'History', 'Year')]
                for i in range(2, sheet.max_row):
                    ID = str(sheet.cell(row = i, column = 1).value)
                    name = str(sheet.cell(row = i, column = 2).value)
                    stn = str(sheet.cell(row = i, column = 3).value)
                    diag = str(sheet.cell(row = i, column = 4).value)
                    tissue = str(sheet.cell(row = i, column = 5).value)
                    hist = str(sheet.cell(row = i, column = 6).value)
                    year = str(sheet.cell(row = i, column = 7).value)
                    slidelist = (ID, name, stn, diag, tissue, hist, year)
                    table.append(slidelist)
                #

                print(tabulate(table))
                confirm = input('Add these. Is this correct? (Y/N): ')
                if confirm == 'Y':
                    addMany(con, cur, xlsxpath)
                    print('Slides added.')
                else:
                    print('Not confirmed. No changes made.')
                #
        elif choice == '7':
            xlsxpath = input('Enter path of .xlsx file: ')
            if os.path.isabs(xlsxpath):

                import openpyxl
                wb = openpyxl.load_workbook(xlsxpath)
                sheet = wb.active

                table = [('Case ID', 'Filename', 'Stain', 'Diagnosis', 'Tissue', 'History', 'Year')]
                for i in range(2, sheet.max_row):
                    ID = str(sheet.cell(row = i, column = 1).value)
                    name = str(sheet.cell(row = i, column = 2).value)
                    stn = str(sheet.cell(row = i, column = 3).value)
                    diag = str(sheet.cell(row = i, column = 4).value)
                    tissue = str(sheet.cell(row = i, column = 5).value)
                    hist = str(sheet.cell(row = i, column = 6).value)
                    year = str(sheet.cell(row = i, column = 7).value)
                    slidelist = (ID, name, stn, diag, tissue, hist, year)
                    table.append(slidelist)
                #

                print(tabulate(table))
                confirm = input('Remove these. Is this correct? (Y/N): ')
                if confirm == 'Y':
                    removeMany(con, cur, xlsxpath)
                    print('Slides removed.')
                else:
                    print('Not confirmed. No changes made.')
                #
            else:
                print('Please enter complete file path.')
            #
        elif choice == '8':
            res = cur.execute("SELECT case_ID, filename, stain, diagnosis, tissue_type, history, year FROM slide ORDER BY case_ID ASC;")
            reslist = res.fetchall()
            table = [('Case ID', 'Filename', 'Stain', 'Diagnosis', 'Tissue', 'History', 'Year')]
            for i in reslist:
                table.append(i)
            #
            xlsxfilename = os.path.join(dirpath, 'slidebox_print.xlsx')
            printDB(table, xlsxfilename)
            print('Database printed to', xlsxfilename)
        else:
            print('Please follow instructions.')
            pass    

except sqlite3.Error as error:
        print('Error occurred - ', error)
#